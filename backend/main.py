from fastapi import FastAPI, Query, Depends, HTTPException, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import base64
import json as _json
from contextlib import asynccontextmanager
from typing import Optional, List
from pydantic import BaseModel
from datetime import datetime, timedelta
from jose import JWTError, jwt
import bcrypt as _bcrypt_lib
import asyncpg
import os
import secrets
from dotenv import load_dotenv

load_dotenv()

# ─── Auth Config ──────────────────────────────────────────────────────────────
SECRET_KEY = os.getenv("SECRET_KEY", "change-this-secret-in-production")
ALGORITHM  = "HS256"
TOKEN_DAYS = 30

bearer = HTTPBearer(auto_error=False)

def hash_pw(pw: str) -> str:
    return _bcrypt_lib.hashpw(pw.encode(), _bcrypt_lib.gensalt()).decode()
def verify_pw(pw: str, h: str) -> bool:
    try:
        return _bcrypt_lib.checkpw(pw.encode(), h.encode())
    except Exception:
        return False
def make_token(payload: dict) -> str:
    p = {**payload, "exp": datetime.utcnow() + timedelta(days=TOKEN_DAYS)}
    return jwt.encode(p, SECRET_KEY, algorithm=ALGORITHM)

async def current_user(cred: HTTPAuthorizationCredentials = Depends(bearer)):
    if not cred:
        raise HTTPException(status_code=401, detail="未登录")
    try:
        return jwt.decode(cred.credentials, SECRET_KEY, algorithms=[ALGORITHM])
    except JWTError:
        raise HTTPException(status_code=401, detail="Token 无效或已过期")

async def admin_required(user=Depends(current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    return user

# ─── Database Pool ────────────────────────────────────────────────────────────
_pool: Optional[asyncpg.Pool] = None

async def get_pool() -> asyncpg.Pool:
    global _pool
    if _pool is None:
        _pool = await asyncpg.create_pool(
            os.getenv("DATABASE_URL"),
            min_size=2,
            max_size=10,
        )
    return _pool

@asynccontextmanager
async def lifespan(app: FastAPI):
    await get_pool()
    yield
    if _pool:
        await _pool.close()

# ─── App ──────────────────────────────────────────────────────────────────────
app = FastAPI(title="GPU·MARKET API", lifespan=lifespan)

CORS_ORIGINS = os.getenv(
    "CORS_ORIGINS",
    "http://localhost:5173,https://www.neocloud.market"
).split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Helpers ──────────────────────────────────────────────────────────────────
def row_to_dict(row) -> dict:
    d = dict(row)
    for k, v in d.items():
        if isinstance(v, (list, tuple)) and not isinstance(v, str):
            d[k] = list(v)
    return d

# ─── Vendors ──────────────────────────────────────────────────────────────────
@app.get("/api/vendors")
async def list_vendors():
    pool = await get_pool()
    sql = """
        SELECT
            v.id,
            v.company_name                          AS name,
            LEFT(v.company_name, 2)                 AS avatar,
            v.rating::float,
            v.review_count                          AS reviews,
            v.location,
            TO_CHAR(v.joined_at, 'YYYY-MM')        AS joined,
            COALESCE(v.slug, v.id::text)            AS "shareToken",
            COALESCE(v.contact_name,'')             AS "contactName",
            COALESCE(v.contact_phone,'')            AS "contactPhone",
            COALESCE(v.email,'')                    AS email
        FROM vendors v
        WHERE v.status = 'active'
        ORDER BY v.rating DESC, v.review_count DESC
    """
    async with pool.acquire() as conn:
        rows = await conn.fetch(sql)
    return [row_to_dict(r) for r in rows]


class VendorCreate(BaseModel):
    name: str
    location: str = "北京"
    contact: str = ""
    phone: str = ""

@app.post("/api/vendors", status_code=201)
async def create_vendor(body: VendorCreate):
    pool = await get_pool()
    sql = """
        INSERT INTO vendors (company_name, location, status, joined_at)
        VALUES ($1, $2, 'active', NOW())
        RETURNING
            id,
            company_name                      AS name,
            LEFT(company_name, 2)             AS avatar,
            rating::float,
            review_count                      AS reviews,
            location,
            TO_CHAR(joined_at, 'YYYY-MM')    AS joined
    """
    async with pool.acquire() as conn:
        row = await conn.fetchrow(sql, body.name, body.location)
    return row_to_dict(row)


# ─── Resources ────────────────────────────────────────────────────────────────
@app.get("/api/resources")
async def list_resources(
    search:    Optional[str]  = Query(None),
    region:    Optional[str]  = Query(None),
    tag:       Optional[str]  = Query(None),
    available: Optional[bool] = Query(None),
):
    pool = await get_pool()

    conditions = ["1=1"]
    params: list = []

    if search:
        params.append(f"%{search}%")
        n = len(params)
        conditions.append(f"(r.gpu_model ILIKE ${n} OR v.company_name ILIKE ${n})")
    if region:
        params.append(region)
        conditions.append(f"r.region = ${len(params)}")
    if available is not None:
        params.append(available)
        conditions.append(f"r.is_available = ${len(params)}")

    where = " AND ".join(conditions)

    # tag filter handled via HAVING after GROUP BY
    having = ""
    if tag:
        params.append(tag)
        having = f"HAVING ${len(params)} = ANY(array_agg(t.name))"

    sql = f"""
        SELECT
            r.id,
            r.vendor_id                                                          AS "vendorId",
            r.gpu_model                                                          AS gpu,
            r.gpu_count                                                          AS count,
            r.price_per_hour::float                                              AS price,
            r.memory_size                                                        AS mem,
            r.memory_bandwidth                                                   AS bandwidth,
            r.is_available                                                       AS available,
            COALESCE(r.resource_status, '在线')                                  AS status,
            r.is_visible                                                         AS "isVisible",
            r.available_quantity                                                 AS "availableQuantity",
            r.delivery_type                                                      AS delivery,
            r.description                                                        AS desc,
            r.region,
            COALESCE(r.billing_unit,'')                                          AS "billingUnit",
            COALESCE(r.contact_name,'')                                          AS "resContactName",
            COALESCE(
                array_agg(t.name ORDER BY t.id) FILTER (WHERE t.name IS NOT NULL),
                ARRAY[]::text[]
            )                                                                    AS tags
        FROM gpu_resources r
        LEFT JOIN vendors  v   ON v.id   = r.vendor_id
        LEFT JOIN resource_tag_map rtm ON rtm.resource_id = r.id
        LEFT JOIN tags     t   ON t.id   = rtm.tag_id
        WHERE {where}
        GROUP BY r.id
        {having}
        ORDER BY r.id
    """
    async with pool.acquire() as conn:
        rows = await conn.fetch(sql, *params)
    return [row_to_dict(r) for r in rows]


class ResourceCreate(BaseModel):
    vendorId: int
    gpu: str
    count: int
    price: float
    mem: str = ""
    bandwidth: str = ""
    region: str = "国内"
    delivery: str = "裸金属"
    desc: str = ""
    tags: List[str] = []
    available: bool = True
    billing_unit: str = ""
    contact_name: Optional[str] = None

@app.post("/api/resources", status_code=201)
async def create_resource(body: ResourceCreate):
    pool = await get_pool()
    async with pool.acquire() as conn:
        # insert resource
        sql = """
            INSERT INTO gpu_resources
                (vendor_id, gpu_model, gpu_count, price_per_hour,
                 memory_size, memory_bandwidth, region,
                 delivery_type, description, is_available,
                 billing_unit, contact_name)
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12)
            RETURNING id
        """
        resource_id = await conn.fetchval(
            sql, body.vendorId, body.gpu, body.count, body.price,
            body.mem, body.bandwidth, body.region,
            body.delivery, body.desc, body.available,
            body.billing_unit or None, body.contact_name or None
        )
        # insert tags
        for tag_name in body.tags:
            tag_id = await conn.fetchval(
                "SELECT id FROM tags WHERE name = $1", tag_name
            )
            if tag_id:
                await conn.execute(
                    "INSERT INTO resource_tag_map (resource_id, tag_id) VALUES ($1,$2) ON CONFLICT DO NOTHING",
                    resource_id, tag_id
                )
    # return full resource
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT r.id, r.vendor_id AS "vendorId", r.gpu_model AS gpu,
                   r.gpu_count AS count, r.price_per_hour::float AS price,
                   r.memory_size AS mem, r.memory_bandwidth AS bandwidth,
                   r.is_available AS available, r.delivery_type AS delivery,
                   r.description AS desc, r.region,
                   COALESCE(array_agg(t.name ORDER BY t.id) FILTER (WHERE t.name IS NOT NULL), ARRAY[]::text[]) AS tags
            FROM gpu_resources r
            LEFT JOIN resource_tag_map rtm ON rtm.resource_id = r.id
            LEFT JOIN tags t ON t.id = rtm.tag_id
            WHERE r.id = $1
            GROUP BY r.id
            """, resource_id
        )
    return row_to_dict(row)


# ─── Demands ──────────────────────────────────────────────────────────────────
@app.get("/api/demands")
async def list_demands():
    pool = await get_pool()
    sql = """
        SELECT
            id, gpu,
            COALESCE(gpu_brand,'')         AS gpu_brand,
            COALESCE(gpu_other,'')         AS gpu_other,
            gpu_count                      AS count,
            COALESCE(count_unit,'卡')      AS count_unit,
            COALESCE(dc_location,'')       AS dc_location,
            COALESCE(rental_months,1)      AS rental_months,
            budget::float, tags, region,
            delivery_type                  AS delivery,
            COALESCE(delivery_other,'')    AS delivery_other,
            COALESCE(contract_type,'')     AS contract_type,
            COALESCE(payment_type,'')      AS payment_type,
            COALESCE(payment_other,'')     AS payment_other,
            COALESCE(config_req,'')        AS config_req,
            COALESCE(storage_req,'')       AS storage_req,
            COALESCE(bandwidth_req,'')     AS bandwidth_req,
            COALESCE(public_ip_req,'')     AS public_ip_req,
            COALESCE(need_extra_cpu,FALSE) AS need_extra_cpu,
            COALESCE(extra_cpu_config,'')  AS extra_cpu_config,
            COALESCE(contact_name,'')      AS contact_name,
            COALESCE(contact_phone,'')     AS contact_phone,
            COALESCE(company,'')           AS company,
            COALESCE(contact_email,'')     AS contact_email,
            COALESCE(notes,'')             AS notes,
            COALESCE(delivery_time,'')     AS delivery_time,
            COALESCE(budget_text,'')       AS budget_text,
            description                    AS desc,
            contact,
            TO_CHAR(created_at,'YYYY-MM-DD') AS "createdAt"
        FROM demands
        WHERE COALESCE(is_visible, TRUE) = TRUE
          AND COALESCE(status, 'online') = 'online'
        ORDER BY created_at DESC
    """
    async with pool.acquire() as conn:
        rows = await conn.fetch(sql)
    return [row_to_dict(r) for r in rows]


class DemandCreate(BaseModel):
    gpu: str
    gpu_brand: str = ""
    gpu_other: str = ""
    count: int
    count_unit: str = "卡"
    dc_location: str = ""
    rental_months: int = 1
    delivery_type: str = "裸金属服务器"
    delivery_other: str = ""
    contract_type: str = "开口合同"
    payment_type: str = "预付"
    payment_other: str = ""
    config_req: str = ""
    storage_req: str = ""
    bandwidth_req: str = ""
    public_ip_req: str = ""
    need_extra_cpu: bool = False
    extra_cpu_config: str = ""
    contact_name: str = ""
    contact_phone: str = ""
    company: str = ""
    contact_email: str = ""
    notes: str = ""
    delivery_time: str = ""
    budget_text: str = ""
    # 旧字段（向后兼容）
    budget: float = 0
    tags: List[str] = []
    region: str = ""
    contact: str = ""
    vendor_id: Optional[int] = None

@app.post("/api/demands", status_code=201)
async def create_demand(body: DemandCreate):
    pool = await get_pool()
    contact_val = body.contact or body.contact_name
    sql = """
        INSERT INTO demands (
            gpu, gpu_brand, gpu_other, gpu_count, count_unit, dc_location, rental_months,
            delivery_type, delivery_other, contract_type, payment_type, payment_other,
            config_req, storage_req, bandwidth_req, public_ip_req,
            need_extra_cpu, extra_cpu_config,
            contact_name, contact_phone, company, contact_email, notes,
            budget, tags, region, description, contact, vendor_id,
            delivery_time, budget_text
        ) VALUES (
            $1,$2,$3,$4,$5,$6,$7,
            $8,$9,$10,$11,$12,
            $13,$14,$15,$16,
            $17,$18,
            $19,$20,$21,$22,$23,
            $24,$25,$26,$27,$28,$29,
            $30,$31
        )
        RETURNING
            id, gpu,
            COALESCE(gpu_brand,'')         AS gpu_brand,
            COALESCE(gpu_other,'')         AS gpu_other,
            gpu_count                      AS count,
            COALESCE(count_unit,'卡')      AS count_unit,
            COALESCE(dc_location,'')       AS dc_location,
            COALESCE(rental_months,1)      AS rental_months,
            budget::float, tags, region,
            delivery_type                  AS delivery,
            COALESCE(delivery_other,'')    AS delivery_other,
            COALESCE(contract_type,'')     AS contract_type,
            COALESCE(payment_type,'')      AS payment_type,
            COALESCE(payment_other,'')     AS payment_other,
            COALESCE(config_req,'')        AS config_req,
            COALESCE(storage_req,'')       AS storage_req,
            COALESCE(bandwidth_req,'')     AS bandwidth_req,
            COALESCE(public_ip_req,'')     AS public_ip_req,
            COALESCE(need_extra_cpu,FALSE) AS need_extra_cpu,
            COALESCE(extra_cpu_config,'')  AS extra_cpu_config,
            COALESCE(contact_name,'')      AS contact_name,
            COALESCE(contact_phone,'')     AS contact_phone,
            COALESCE(company,'')           AS company,
            COALESCE(contact_email,'')     AS contact_email,
            COALESCE(notes,'')             AS notes,
            COALESCE(delivery_time,'')     AS delivery_time,
            COALESCE(budget_text,'')       AS budget_text,
            description                    AS desc,
            contact,
            TO_CHAR(created_at,'YYYY-MM-DD') AS "createdAt"
    """
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            sql,
            body.gpu, body.gpu_brand, body.gpu_other, body.count, body.count_unit, body.dc_location, body.rental_months,
            body.delivery_type, body.delivery_other, body.contract_type, body.payment_type, body.payment_other,
            body.config_req, body.storage_req, body.bandwidth_req, body.public_ip_req,
            body.need_extra_cpu, body.extra_cpu_config,
            body.contact_name, body.contact_phone, body.company, body.contact_email, body.notes,
            body.budget, body.tags, body.region, body.notes, contact_val, body.vendor_id,
            body.delivery_time or None, body.budget_text or None
        )
    return row_to_dict(row)


# ─── Subscribers ──────────────────────────────────────────────────────────────
@app.get("/api/subscribers/count")
async def subscriber_count():
    pool = await get_pool()
    sql = """
        SELECT
            COUNT(*) FILTER (WHERE 'resources' = ANY(topics)) AS resources,
            COUNT(*) FILTER (WHERE 'demands'   = ANY(topics)) AS demands
        FROM subscribers
    """
    async with pool.acquire() as conn:
        row = await conn.fetchrow(sql)
    return {"resources": int(row["resources"]), "demands": int(row["demands"])}


class SubscriberCreate(BaseModel):
    email: str
    topics: List[str]

@app.post("/api/subscribers", status_code=201)
async def create_subscriber(body: SubscriberCreate):
    pool = await get_pool()
    sql = """
        INSERT INTO subscribers (email, topics)
        VALUES ($1, $2)
        ON CONFLICT (email) DO UPDATE SET topics = EXCLUDED.topics
        RETURNING id, email, topics
    """
    async with pool.acquire() as conn:
        row = await conn.fetchrow(sql, body.email, body.topics)
    return row_to_dict(row)


# ─── Auth ──────────────────────────────────────────────────────────────────────
VENDOR_SELECT = """
    SELECT id, company_name AS name, LEFT(company_name,2) AS avatar,
           rating::float, review_count AS reviews, location,
           TO_CHAR(joined_at,'YYYY-MM') AS joined, slug,
           contact_name AS "contactName", contact_phone AS "contactPhone", email
    FROM vendors WHERE user_id = $1
"""

class RegisterBody(BaseModel):
    email: str
    password: str
    company_name: str
    location: str = "北京"
    contact_name: str = ""
    phone: str = ""

@app.post("/auth/register", status_code=201)
async def register(body: RegisterBody):
    pool = await get_pool()
    async with pool.acquire() as conn:
        if await conn.fetchval("SELECT id FROM users WHERE email=$1", body.email):
            raise HTTPException(status_code=400, detail="该邮箱已注册")
        username = body.email.split("@")[0] + "_" + str(int(datetime.utcnow().timestamp()))[-5:]
        user_id = await conn.fetchval("""
            INSERT INTO users (email, phone, password_hash, username, display_name, role, status, email_verified_at)
            VALUES ($1, NULLIF($2,''), $3, $4, $5, 'vendor', 'active', NOW())
            RETURNING id
        """, body.email, body.phone, hash_pw(body.password), username, body.contact_name or body.company_name)
        vendor = await conn.fetchrow("""
            INSERT INTO vendors (user_id, company_name, location, contact_name, contact_phone, email, status, joined_at)
            VALUES ($1, $2, $3, $4, $5, $6, 'active', NOW())
            RETURNING id, company_name AS name, LEFT(company_name,2) AS avatar,
                      rating::float, review_count AS reviews, location,
                      TO_CHAR(joined_at,'YYYY-MM') AS joined
        """, user_id, body.company_name, body.location,
             body.contact_name or None, body.phone or None, body.email)
    v = row_to_dict(vendor)
    token = make_token({"sub": str(user_id), "vendor_id": v["id"]})
    return {"token": token, "vendor": v}


class LoginBody(BaseModel):
    email: str
    password: str

@app.post("/auth/login")
async def login(body: LoginBody):
    pool = await get_pool()
    async with pool.acquire() as conn:
        user = await conn.fetchrow(
            "SELECT id, password_hash, role FROM users WHERE email=$1 OR phone=$1", body.email
        )
        if not user or not verify_pw(body.password, user["password_hash"]):
            raise HTTPException(status_code=401, detail="邮箱或密码错误")
        if user["role"] == "admin":
            token = make_token({"sub": str(user["id"]), "role": "admin"})
            return {"token": token, "role": "admin"}
        vendor = await conn.fetchrow(VENDOR_SELECT, user["id"])
    if not vendor:
        raise HTTPException(status_code=403, detail="该账号不是供应商账号")
    v = row_to_dict(vendor)
    token = make_token({"sub": str(user["id"]), "vendor_id": v["id"], "role": "vendor"})
    return {"token": token, "role": "vendor", "vendor": v}


class ResourcePatch(BaseModel):
    status: Optional[str] = None
    count: Optional[int] = None
    is_visible: Optional[bool] = None
    available_quantity: Optional[int] = None

@app.patch("/api/resources/{resource_id}", status_code=200)
async def patch_resource(resource_id: int, body: ResourcePatch, user=Depends(current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT vendor_id FROM gpu_resources WHERE id=$1", resource_id)
        if not row or row["vendor_id"] != user["vendor_id"]:
            raise HTTPException(status_code=403, detail="无权操作")
        updates = []
        params = []
        if body.status is not None:
            params.append(body.status)
            updates.append(f"resource_status=${len(params)}")
            params.append(body.status != "下架")
            updates.append(f"is_available=${len(params)}")
        if body.count is not None:
            params.append(body.count)
            updates.append(f"gpu_count=${len(params)}")
        if body.is_visible is not None:
            params.append(body.is_visible)
            updates.append(f"is_visible=${len(params)}")
        if body.available_quantity is not None:
            params.append(body.available_quantity)
            updates.append(f"available_quantity=${len(params)}")
        if not updates:
            return {"ok": True}
        params.append(resource_id)
        await conn.execute(f"UPDATE gpu_resources SET {','.join(updates)} WHERE id=${len(params)}", *params)
    return {"ok": True}


@app.delete("/api/resources/{resource_id}", status_code=204)
async def delete_resource(resource_id: int, user=Depends(current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT vendor_id FROM gpu_resources WHERE id=$1", resource_id)
        if not row or row["vendor_id"] != user["vendor_id"]:
            raise HTTPException(status_code=403, detail="无权操作")
        await conn.execute("DELETE FROM gpu_resources WHERE id=$1", resource_id)


class VendorSlug(BaseModel):
    slug: str

class VendorContact(BaseModel):
    contact_name: str = ""
    contact_phone: str = ""
    email: str = ""

@app.patch("/api/vendors/contact")
async def update_vendor_contact(body: VendorContact, user=Depends(current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE vendors SET contact_name=$1, contact_phone=$2, email=$3 WHERE id=$4",
            body.contact_name or None, body.contact_phone or None, body.email or None,
            user["vendor_id"]
        )
    return {"ok": True}

@app.patch("/api/vendors/slug")
async def update_vendor_slug(body: VendorSlug, user=Depends(current_user)):
    import re
    if not re.match(r'^[a-zA-Z0-9\-]{2,40}$', body.slug):
        raise HTTPException(status_code=400, detail="slug 只能包含字母、数字、连字符，长度 2-40")
    pool = await get_pool()
    async with pool.acquire() as conn:
        conflict = await conn.fetchval(
            "SELECT id FROM vendors WHERE slug=$1 AND id!=$2", body.slug, user["vendor_id"]
        )
        if conflict:
            raise HTTPException(status_code=409, detail="该名称已被使用，请换一个")
        await conn.execute("UPDATE vendors SET slug=$1 WHERE id=$2", body.slug, user["vendor_id"])
    return {"slug": body.slug}

@app.post("/api/vendors/share-token")
async def get_or_create_share_token(user=Depends(current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT slug FROM vendors WHERE id=$1", user["vendor_id"])
    token = row["slug"] if row and row["slug"] else str(user["vendor_id"])
    return {"share_token": token}


@app.get("/api/share/{share_token}")
async def get_share_page(share_token: str):
    pool = await get_pool()
    async with pool.acquire() as conn:
        # 先按 slug 查，查不到再按 ID 查
        vendor = await conn.fetchrow("""
            SELECT id, company_name AS name, LEFT(company_name,2) AS avatar,
                   rating::float, review_count AS reviews, location,
                   TO_CHAR(joined_at,'YYYY-MM') AS joined,
                   contact_name AS "contactName", contact_phone AS "contactPhone", email
            FROM vendors WHERE slug=$1 AND status='active'
        """, share_token)
        if not vendor:
            try:
                vendor_id = int(share_token)
            except ValueError:
                raise HTTPException(status_code=404, detail="分享链接无效")
            vendor = await conn.fetchrow("""
                SELECT id, company_name AS name, LEFT(company_name,2) AS avatar,
                       rating::float, review_count AS reviews, location,
                       TO_CHAR(joined_at,'YYYY-MM') AS joined,
                       contact_name AS "contactName", contact_phone AS "contactPhone", email
                FROM vendors WHERE id=$1 AND status='active'
            """, vendor_id)
        if not vendor:
            raise HTTPException(status_code=404, detail="分享链接无效")
        vendor_dict = row_to_dict(vendor)
        resources = await conn.fetch("""
            SELECT r.id, r.vendor_id AS "vendorId", r.gpu_model AS gpu,
                   r.gpu_count AS count, r.price_per_hour::float AS price,
                   r.memory_size AS mem, r.memory_bandwidth AS bandwidth,
                   r.is_available AS available,
                   COALESCE(r.resource_status, '在线') AS status,
                   r.is_visible AS "isVisible",
                   r.available_quantity AS "availableQuantity",
                   r.delivery_type AS delivery, r.description AS desc, r.region,
                   COALESCE(
                       array_agg(t.name ORDER BY t.id) FILTER (WHERE t.name IS NOT NULL),
                       ARRAY[]::text[]
                   ) AS tags
            FROM gpu_resources r
            LEFT JOIN resource_tag_map rtm ON rtm.resource_id = r.id
            LEFT JOIN tags t ON t.id = rtm.tag_id
            WHERE r.vendor_id=$1 AND r.is_visible = TRUE
            GROUP BY r.id
            ORDER BY r.id
        """, vendor_dict["id"])
    return {"vendor": vendor_dict, "resources": [row_to_dict(r) for r in resources]}


# ─── Related Links ────────────────────────────────────────────────────────────
class RelatedLinkCreate(BaseModel):
    title: str
    url: str
    sort_order: int = 0

@app.get("/api/related-links")
async def list_related_links():
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, title, url, sort_order FROM related_links ORDER BY sort_order, id"
        )
    return [row_to_dict(r) for r in rows]

@app.post("/api/related-links", status_code=201)
async def create_related_link(body: RelatedLinkCreate, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "INSERT INTO related_links (title, url, sort_order) VALUES ($1, $2, $3) RETURNING id, title, url, sort_order",
            body.title, body.url, body.sort_order
        )
    return row_to_dict(row)

@app.delete("/api/related-links/{link_id}", status_code=204)
async def delete_related_link(link_id: int, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM related_links WHERE id=$1", link_id)


# ─── Admin Vendor Management ──────────────────────────────────────────────────
@app.get("/api/admin/vendors")
async def admin_list_vendors(user=Depends(admin_required)):
    pool = await get_pool()
    sql = """
        SELECT
            v.id,
            v.company_name                      AS name,
            COALESCE(v.contact_name, '')        AS contact_name,
            COALESCE(v.contact_phone, '')       AS contact_phone,
            COALESCE(v.email, '')               AS email,
            v.location,
            v.status,
            TO_CHAR(v.joined_at, 'YYYY-MM-DD') AS joined,
            COUNT(r.id)::int                    AS resource_count
        FROM vendors v
        LEFT JOIN gpu_resources r ON r.vendor_id = v.id
        GROUP BY v.id
        ORDER BY v.joined_at DESC
    """
    async with pool.acquire() as conn:
        rows = await conn.fetch(sql)
    return [row_to_dict(r) for r in rows]


class VendorStatusPatch(BaseModel):
    status: str  # 'active' | 'suspended'

@app.patch("/api/admin/vendors/{vendor_id}")
async def admin_patch_vendor(vendor_id: int, body: VendorStatusPatch, user=Depends(admin_required)):
    if body.status not in ("active", "suspended"):
        raise HTTPException(status_code=400, detail="status 只能为 active 或 suspended")
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT id FROM vendors WHERE id=$1", vendor_id)
        if not row:
            raise HTTPException(status_code=404, detail="供应商不存在")
        await conn.execute("UPDATE vendors SET status=$1 WHERE id=$2", body.status, vendor_id)
    return {"ok": True, "status": body.status}


# ─── GPU Model Library ────────────────────────────────────────────────────────
class GpuBrandCreate(BaseModel):
    name: str
    logo_url: str = ""
    sort_order: int = 0

class GpuSeriesCreate(BaseModel):
    brand_id: int
    name: str
    sort_order: int = 0

class GpuModelCreate(BaseModel):
    series_id: int
    brand_id: int
    name: str
    vram_gb: Optional[float] = None
    tdp_w: Optional[int] = None
    architecture: str = ""
    sort_order: int = 0
    is_active: bool = True

class GpuModelPatch(BaseModel):
    name: Optional[str] = None
    vram_gb: Optional[float] = None
    architecture: Optional[str] = None
    sort_order: Optional[int] = None
    is_active: Optional[bool] = None

@app.get("/api/gpu-brands")
async def list_gpu_brands():
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT id, name, logo_url, sort_order FROM gpu_brands ORDER BY sort_order, name")
    return [row_to_dict(r) for r in rows]

@app.get("/api/gpu-series")
async def list_gpu_series(brand_id: int = Query(...)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, brand_id, name, sort_order FROM gpu_series WHERE brand_id=$1 ORDER BY sort_order, name",
            brand_id)
    return [row_to_dict(r) for r in rows]

@app.get("/api/gpu-models")
async def list_gpu_models(series_id: int = Query(...), include_inactive: bool = Query(False)):
    pool = await get_pool()
    sql = "SELECT id, series_id, brand_id, name, vram_gb::float, tdp_w, architecture, sort_order, is_active FROM gpu_models WHERE series_id=$1"
    if not include_inactive:
        sql += " AND is_active=TRUE"
    sql += " ORDER BY sort_order, name"
    async with pool.acquire() as conn:
        rows = await conn.fetch(sql, series_id)
    return [row_to_dict(r) for r in rows]

@app.post("/api/admin/gpu-brands", status_code=201)
async def admin_create_gpu_brand(body: GpuBrandCreate, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "INSERT INTO gpu_brands (name, logo_url, sort_order) VALUES ($1, $2, $3) RETURNING id, name, logo_url, sort_order",
            body.name, body.logo_url or None, body.sort_order)
    return row_to_dict(row)

@app.delete("/api/admin/gpu-brands/{brand_id}", status_code=204)
async def admin_delete_gpu_brand(brand_id: int, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM gpu_brands WHERE id=$1", brand_id)

@app.post("/api/admin/gpu-series", status_code=201)
async def admin_create_gpu_series(body: GpuSeriesCreate, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "INSERT INTO gpu_series (brand_id, name, sort_order) VALUES ($1, $2, $3) RETURNING id, brand_id, name, sort_order",
            body.brand_id, body.name, body.sort_order)
    return row_to_dict(row)

@app.delete("/api/admin/gpu-series/{series_id}", status_code=204)
async def admin_delete_gpu_series(series_id: int, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM gpu_series WHERE id=$1", series_id)

@app.post("/api/admin/gpu-models", status_code=201)
async def admin_create_gpu_model(body: GpuModelCreate, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "INSERT INTO gpu_models (series_id,brand_id,name,vram_gb,tdp_w,architecture,sort_order,is_active) "
            "VALUES ($1,$2,$3,$4,$5,$6,$7,$8) "
            "RETURNING id, series_id, brand_id, name, vram_gb::float, tdp_w, architecture, sort_order, is_active",
            body.series_id, body.brand_id, body.name, body.vram_gb, body.tdp_w,
            body.architecture or None, body.sort_order, body.is_active)
    return row_to_dict(row)

@app.patch("/api/admin/gpu-models/{model_id}")
async def admin_patch_gpu_model(model_id: int, body: GpuModelPatch, user=Depends(admin_required)):
    updates = {k: v for k, v in body.dict().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="没有字段需要更新")
    pool = await get_pool()
    set_clauses = ", ".join(f"{k}=${i+2}" for i, k in enumerate(updates.keys()))
    async with pool.acquire() as conn:
        await conn.execute(f"UPDATE gpu_models SET {set_clauses} WHERE id=$1", model_id, *updates.values())
    return {"ok": True}

@app.delete("/api/admin/gpu-models/{model_id}", status_code=204)
async def admin_delete_gpu_model(model_id: int, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM gpu_models WHERE id=$1", model_id)


EXTRACT_SYSTEM = """你是一个 GPU 租赁信息提取助手。从用户上传的文件或截图中提取以下字段，返回 JSON 格式，没有找到的字段返回 null：
{"gpu_model":"GPU型号","quantity":"数量（整数）","quantity_unit":"台或卡","location":"机房位置","rental_period":"租赁周期（月数整数）","delivery_type":"交付形式","contract_type":"合同形式","payment_type":"付款方式","config_requirements":"配置要求","storage_requirements":"存储要求","bandwidth_requirements":"带宽要求","ip_requirements":"公网IP要求","extra_cpu":"是否需要额外CPU（true/false）","extra_cpu_config":"CPU补充配置","contact_name":"联系人姓名","contact_phone":"联系电话","company_name":"公司名称","email":"电子邮箱","remarks":"备注"}
只返回 JSON，不要任何解释。"""

@app.post("/api/extract-info")
async def extract_info(file: UploadFile = File(...)):
    import anthropic
    content = await file.read()
    ct = (file.content_type or "").lower()
    client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

    if ct in ("image/png", "image/jpeg", "image/jpg", "image/webp"):
        b64 = base64.standard_b64encode(content).decode()
        media = "image/jpeg" if ct == "image/jpg" else ct
        msg = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=1024, system=EXTRACT_SYSTEM,
            messages=[{"role":"user","content":[
                {"type":"image","source":{"type":"base64","media_type":media,"data":b64}},
                {"type":"text","text":"请提取信息"}
            ]}]
        )
    elif ct == "application/pdf":
        b64 = base64.standard_b64encode(content).decode()
        msg = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=1024, system=EXTRACT_SYSTEM,
            messages=[{"role":"user","content":[
                {"type":"document","source":{"type":"base64","media_type":"application/pdf","data":b64}},
                {"type":"text","text":"请提取信息"}
            ]}]
        )
    elif "wordprocessingml" in ct:
        import io
        from docx import Document
        doc = Document(io.BytesIO(content))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        msg = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=1024, system=EXTRACT_SYSTEM,
            messages=[{"role":"user","content":text}]
        )
    elif "spreadsheetml" in ct:
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                r = "\t".join(str(v) if v is not None else "" for v in row)
                if r.strip(): lines.append(r)
        msg = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=1024, system=EXTRACT_SYSTEM,
            messages=[{"role":"user","content":"\n".join(lines[:200])}]
        )
    else:
        raise HTTPException(status_code=400, detail="不支持的文件格式，请上传 PNG/JPG/PDF/DOCX/XLSX")

    raw = msg.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"): raw = raw[4:]
    return _json.loads(raw.strip())


# ─── Admin Extended Endpoints ─────────────────────────────────────────────────
class VendorInfoPatch(BaseModel):
    company_name: str = ""
    contact_name: str = ""
    contact_phone: str = ""
    email: str = ""
    location: str = ""

@app.patch("/api/admin/vendors/{vendor_id}/info")
async def admin_patch_vendor_info(vendor_id: int, body: VendorInfoPatch, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT id FROM vendors WHERE id=$1", vendor_id)
        if not row:
            raise HTTPException(status_code=404, detail="供应商不存在")
        await conn.execute(
            "UPDATE vendors SET company_name=COALESCE(NULLIF($1,''),company_name), contact_name=COALESCE(NULLIF($2,''),contact_name), contact_phone=COALESCE(NULLIF($3,''),contact_phone), email=COALESCE(NULLIF($4,''),email), location=COALESCE(NULLIF($5,''),location) WHERE id=$6",
            body.company_name, body.contact_name, body.contact_phone, body.email, body.location, vendor_id
        )
    return {"ok": True}

@app.get("/api/admin/vendors/{vendor_id}/resources")
async def admin_vendor_resources(vendor_id: int, user=Depends(admin_required)):
    pool = await get_pool()
    sql = """
        SELECT r.id, r.gpu_model AS gpu, r.gpu_count AS count,
               r.price_per_hour::float AS price,
               COALESCE(r.resource_status, '在线') AS status,
               COALESCE(r.is_visible, TRUE) AS is_visible
        FROM gpu_resources r
        WHERE r.vendor_id = $1
        ORDER BY r.id DESC
    """
    async with pool.acquire() as conn:
        rows = await conn.fetch(sql, vendor_id)
    return [row_to_dict(r) for r in rows]


class VisibilityPatch(BaseModel):
    is_visible: bool

@app.patch("/api/admin/resources/{resource_id}/visibility")
async def admin_patch_resource_visibility(resource_id: int, body: VisibilityPatch, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("UPDATE gpu_resources SET is_visible=$1 WHERE id=$2", body.is_visible, resource_id)
    return {"ok": True}

@app.delete("/api/admin/resources/{resource_id}", status_code=204)
async def admin_delete_resource(resource_id: int, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM gpu_resources WHERE id=$1", resource_id)

@app.get("/api/admin/vendors/{vendor_id}/demands")
async def admin_vendor_demands(vendor_id: int, user=Depends(admin_required)):
    pool = await get_pool()
    sql = """
        SELECT id, gpu, gpu_count AS count, COALESCE(count_unit,'卡') AS count_unit,
               COALESCE(contact_name,'') AS contact_name,
               TO_CHAR(created_at,'YYYY-MM-DD') AS "createdAt",
               COALESCE(is_visible, TRUE) AS is_visible
        FROM demands
        WHERE vendor_id = $1
        ORDER BY created_at DESC
    """
    async with pool.acquire() as conn:
        rows = await conn.fetch(sql, vendor_id)
    return [row_to_dict(r) for r in rows]

@app.patch("/api/admin/demands/{demand_id}/visibility")
async def admin_patch_demand_visibility(demand_id: int, body: VisibilityPatch, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("UPDATE demands SET is_visible=$1 WHERE id=$2", body.is_visible, demand_id)
    return {"ok": True}

@app.delete("/api/admin/demands/{demand_id}", status_code=204)
async def admin_delete_demand(demand_id: int, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM demands WHERE id=$1", demand_id)

@app.get("/api/admin/demands")
async def admin_all_demands(user=Depends(admin_required)):
    pool = await get_pool()
    sql = """
        SELECT d.id, d.gpu, COALESCE(d.gpu_brand,'') AS gpu_brand,
               d.gpu_count AS count, COALESCE(d.count_unit,'卡') AS count_unit,
               COALESCE(d.rental_months,0) AS rental_months,
               COALESCE(d.region,'') AS region,
               COALESCE(d.dc_location,'') AS dc_location,
               COALESCE(d.delivery_type,'') AS delivery_type,
               COALESCE(d.contract_type,'') AS contract_type,
               COALESCE(d.payment_type,'') AS payment_type,
               COALESCE(d.budget_text,'') AS budget_text,
               d.budget::float AS budget,
               COALESCE(d.delivery_time,'') AS delivery_time,
               COALESCE(d.config_req,'') AS config_req,
               COALESCE(d.notes,'') AS notes,
               COALESCE(d.contact_name,'') AS contact_name,
               COALESCE(d.contact_phone,'') AS contact_phone,
               COALESCE(d.company,'') AS company,
               COALESCE(d.contact_email,'') AS contact_email,
               COALESCE(d.is_visible, TRUE) AS is_visible,
               d.vendor_id,
               COALESCE(v.company_name,'') AS vendor_name,
               TO_CHAR(d.created_at,'YYYY-MM-DD') AS "createdAt"
        FROM demands d
        LEFT JOIN vendors v ON v.id = d.vendor_id
        ORDER BY d.created_at DESC
    """
    async with pool.acquire() as conn:
        rows = await conn.fetch(sql)
    return [row_to_dict(r) for r in rows]

class DemandPatch(BaseModel):
    gpu: Optional[str] = None
    gpu_brand: Optional[str] = None
    gpu_count: Optional[int] = None
    count_unit: Optional[str] = None
    rental_months: Optional[int] = None
    region: Optional[str] = None
    dc_location: Optional[str] = None
    delivery_time: Optional[str] = None
    contract_type: Optional[str] = None
    payment_type: Optional[str] = None
    budget_text: Optional[str] = None
    notes: Optional[str] = None
    is_visible: Optional[bool] = None
    vendor_id: Optional[int] = None

@app.patch("/api/admin/demands/{demand_id}")
async def admin_patch_demand(demand_id: int, body: DemandPatch, user=Depends(admin_required)):
    pool = await get_pool()
    updates = []
    params: list = []
    i = 1
    for field, col in [
        ("gpu","gpu"), ("gpu_brand","gpu_brand"), ("gpu_count","gpu_count"),
        ("count_unit","count_unit"), ("rental_months","rental_months"),
        ("region","region"), ("dc_location","dc_location"),
        ("delivery_time","delivery_time"), ("contract_type","contract_type"),
        ("payment_type","payment_type"), ("budget_text","budget_text"),
        ("notes","notes"), ("is_visible","is_visible"),
    ]:
        val = getattr(body, field)
        if val is not None:
            updates.append(f"{col}=${i}"); params.append(val); i += 1
    # vendor_id handled separately (allow setting to None via explicit null)
    if body.vendor_id is not None:
        updates.append(f"vendor_id=${i}"); params.append(body.vendor_id); i += 1
    if not updates:
        return {"ok": True}
    async with pool.acquire() as conn:
        await conn.execute(
            f"UPDATE demands SET {', '.join(updates)} WHERE id=${i}",
            *params, demand_id
        )
    return {"ok": True}


# ─── Vendor Demand Endpoints ──────────────────────────────────────────────────
@app.get("/api/vendor/demands")
async def vendor_my_demands(user=Depends(current_user)):
    import uuid
    pool = await get_pool()
    async with pool.acquire() as conn:
        vendor = await conn.fetchrow("SELECT id FROM vendors WHERE user_id=$1", uuid.UUID(user["sub"]))
        if not vendor:
            raise HTTPException(status_code=404, detail="供应商信息不存在")
        rows = await conn.fetch("""
            SELECT id, gpu, COALESCE(gpu_brand,'') AS gpu_brand,
                   COALESCE(gpu_other,'') AS gpu_other,
                   gpu_count AS count, COALESCE(count_unit,'卡') AS count_unit,
                   COALESCE(rental_months,0) AS rental_months,
                   COALESCE(region,'') AS region,
                   COALESCE(dc_location,'') AS dc_location,
                   COALESCE(delivery_type,'') AS delivery_type,
                   COALESCE(delivery_other,'') AS delivery_other,
                   COALESCE(contract_type,'') AS contract_type,
                   COALESCE(payment_type,'') AS payment_type,
                   COALESCE(payment_other,'') AS payment_other,
                   COALESCE(config_req,'') AS config_req,
                   COALESCE(storage_req,'') AS storage_req,
                   COALESCE(bandwidth_req,'') AS bandwidth_req,
                   COALESCE(public_ip_req,'') AS public_ip_req,
                   COALESCE(need_extra_cpu, FALSE) AS need_extra_cpu,
                   COALESCE(extra_cpu_config,'') AS extra_cpu_config,
                   COALESCE(contact_name,'') AS contact_name,
                   COALESCE(contact_phone,'') AS contact_phone,
                   COALESCE(company,'') AS company,
                   COALESCE(contact_email,'') AS contact_email,
                   COALESCE(budget_text,'') AS budget_text,
                   COALESCE(notes,'') AS notes,
                   COALESCE(delivery_time,'') AS delivery_time,
                   COALESCE(status,'online') AS status,
                   COALESCE(is_visible, TRUE) AS is_visible,
                   TO_CHAR(created_at,'YYYY-MM-DD') AS "createdAt"
            FROM demands WHERE vendor_id=$1
            ORDER BY created_at DESC
        """, vendor["id"])
    return [row_to_dict(r) for r in rows]

class VendorDemandUpdate(BaseModel):
    gpu: Optional[str] = None
    gpu_brand: Optional[str] = None
    gpu_count: Optional[int] = None
    count_unit: Optional[str] = None
    rental_months: Optional[int] = None
    region: Optional[str] = None
    dc_location: Optional[str] = None
    delivery_time: Optional[str] = None
    delivery_type: Optional[str] = None
    delivery_other: Optional[str] = None
    contract_type: Optional[str] = None
    payment_type: Optional[str] = None
    payment_other: Optional[str] = None
    config_req: Optional[str] = None
    storage_req: Optional[str] = None
    bandwidth_req: Optional[str] = None
    public_ip_req: Optional[str] = None
    need_extra_cpu: Optional[bool] = None
    extra_cpu_config: Optional[str] = None
    contact_name: Optional[str] = None
    contact_phone: Optional[str] = None
    company: Optional[str] = None
    contact_email: Optional[str] = None
    budget_text: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None

@app.patch("/api/demands/{demand_id}")
async def vendor_patch_demand(demand_id: int, body: VendorDemandUpdate, user=Depends(current_user)):
    import uuid
    pool = await get_pool()
    async with pool.acquire() as conn:
        vendor = await conn.fetchrow("SELECT id FROM vendors WHERE user_id=$1", uuid.UUID(user["sub"]))
        if not vendor:
            raise HTTPException(status_code=403, detail="无供应商权限")
        row = await conn.fetchrow("SELECT vendor_id FROM demands WHERE id=$1", demand_id)
        if not row or row["vendor_id"] != vendor["id"]:
            raise HTTPException(status_code=403, detail="无权限修改此需求")
        fields = [
            ("gpu","gpu"), ("gpu_brand","gpu_brand"), ("gpu_count","gpu_count"),
            ("count_unit","count_unit"), ("rental_months","rental_months"),
            ("region","region"), ("dc_location","dc_location"),
            ("delivery_time","delivery_time"), ("delivery_type","delivery_type"),
            ("delivery_other","delivery_other"), ("contract_type","contract_type"),
            ("payment_type","payment_type"), ("payment_other","payment_other"),
            ("config_req","config_req"), ("storage_req","storage_req"),
            ("bandwidth_req","bandwidth_req"), ("public_ip_req","public_ip_req"),
            ("need_extra_cpu","need_extra_cpu"), ("extra_cpu_config","extra_cpu_config"),
            ("contact_name","contact_name"), ("contact_phone","contact_phone"),
            ("company","company"), ("contact_email","contact_email"),
            ("budget_text","budget_text"), ("notes","notes"), ("status","status"),
        ]
        updates, params, i = [], [], 1
        for field, col in fields:
            val = getattr(body, field)
            if val is not None:
                updates.append(f"{col}=${i}"); params.append(val); i += 1
        if updates:
            await conn.execute(
                f"UPDATE demands SET {', '.join(updates)} WHERE id=${i}",
                *params, demand_id
            )
    return {"ok": True}

@app.delete("/api/demands/{demand_id}")
async def vendor_delete_demand(demand_id: int, user=Depends(current_user)):
    import uuid
    pool = await get_pool()
    async with pool.acquire() as conn:
        vendor = await conn.fetchrow("SELECT id FROM vendors WHERE user_id=$1", uuid.UUID(user["sub"]))
        if not vendor:
            raise HTTPException(status_code=403, detail="无供应商权限")
        row = await conn.fetchrow("SELECT vendor_id FROM demands WHERE id=$1", demand_id)
        if not row or row["vendor_id"] != vendor["id"]:
            raise HTTPException(status_code=403, detail="无权限删除此需求")
        await conn.execute("DELETE FROM demands WHERE id=$1", demand_id)
    return {"ok": True}


# ─── Docs ─────────────────────────────────────────────────────────────────────
import uuid as _uuid_mod

class DocCreate(BaseModel):
    category: str
    title: str
    content: str = ""
    sort_order: int = 0
    is_published: bool = True

class DocPatch(BaseModel):
    category: Optional[str] = None
    title: Optional[str] = None
    content: Optional[str] = None
    sort_order: Optional[int] = None
    is_published: Optional[bool] = None

class DocCommentCreate(BaseModel):
    doc_id: str
    nickname: str = ""
    content: str

@app.get("/api/docs")
async def list_docs():
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, category, title, content, sort_order, is_published, created_at, updated_at "
            "FROM docs WHERE is_published=TRUE ORDER BY sort_order, created_at"
        )
    result = []
    for r in rows:
        d = dict(r)
        d["id"] = str(d["id"])
        d["created_at"] = d["created_at"].isoformat() if d["created_at"] else None
        d["updated_at"] = d["updated_at"].isoformat() if d["updated_at"] else None
        result.append(d)
    return result

@app.get("/api/admin/docs")
async def admin_list_docs(user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, category, title, content, sort_order, is_published, created_at, updated_at "
            "FROM docs ORDER BY sort_order, created_at"
        )
    result = []
    for r in rows:
        d = dict(r)
        d["id"] = str(d["id"])
        d["created_at"] = d["created_at"].isoformat() if d["created_at"] else None
        d["updated_at"] = d["updated_at"].isoformat() if d["updated_at"] else None
        result.append(d)
    return result

@app.post("/api/admin/docs", status_code=201)
async def admin_create_doc(body: DocCreate, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "INSERT INTO docs (category, title, content, sort_order, is_published) "
            "VALUES ($1,$2,$3,$4,$5) RETURNING id, category, title, content, sort_order, is_published, created_at, updated_at",
            body.category, body.title, body.content, body.sort_order, body.is_published
        )
    d = dict(row)
    d["id"] = str(d["id"])
    d["created_at"] = d["created_at"].isoformat() if d["created_at"] else None
    d["updated_at"] = d["updated_at"].isoformat() if d["updated_at"] else None
    return d

@app.patch("/api/admin/docs/{doc_id}")
async def admin_patch_doc(doc_id: str, body: DocPatch, user=Depends(admin_required)):
    pool = await get_pool()
    updates, params, i = [], [], 1
    for field, col in [("category","category"),("title","title"),("content","content"),
                       ("sort_order","sort_order"),("is_published","is_published")]:
        val = getattr(body, field)
        if val is not None:
            updates.append(f"{col}=${i}"); params.append(val); i += 1
    if not updates:
        return {"ok": True}
    updates.append(f"updated_at=NOW()")
    async with pool.acquire() as conn:
        await conn.execute(
            f"UPDATE docs SET {', '.join(updates)} WHERE id=${i}",
            *params, _uuid_mod.UUID(doc_id)
        )
    return {"ok": True}

@app.delete("/api/admin/docs/{doc_id}", status_code=204)
async def admin_delete_doc(doc_id: str, user=Depends(admin_required)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM docs WHERE id=$1", _uuid_mod.UUID(doc_id))

@app.get("/api/docs/{doc_id}/comments")
async def list_doc_comments(doc_id: str):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, doc_id, user_id, nickname, content, created_at "
            "FROM doc_comments WHERE doc_id=$1 ORDER BY created_at",
            _uuid_mod.UUID(doc_id)
        )
    result = []
    for r in rows:
        d = dict(r)
        d["id"] = str(d["id"])
        d["doc_id"] = str(d["doc_id"])
        d["user_id"] = str(d["user_id"]) if d["user_id"] else None
        d["created_at"] = d["created_at"].isoformat() if d["created_at"] else None
        result.append(d)
    return result

@app.post("/api/docs/{doc_id}/comments", status_code=201)
async def create_doc_comment(doc_id: str, body: DocCommentCreate, user=Depends(current_user)):
    pool = await get_pool()
    user_uuid = _uuid_mod.UUID(user["sub"])
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "INSERT INTO doc_comments (doc_id, user_id, nickname, content) VALUES ($1,$2,$3,$4) "
            "RETURNING id, doc_id, user_id, nickname, content, created_at",
            _uuid_mod.UUID(doc_id), user_uuid, body.nickname or None, body.content
        )
    d = dict(row)
    d["id"] = str(d["id"])
    d["doc_id"] = str(d["doc_id"])
    d["user_id"] = str(d["user_id"]) if d["user_id"] else None
    d["created_at"] = d["created_at"].isoformat() if d["created_at"] else None
    return d

@app.delete("/api/docs/comments/{comment_id}", status_code=204)
async def delete_doc_comment(comment_id: str, user=Depends(current_user)):
    pool = await get_pool()
    user_uuid = _uuid_mod.UUID(user["sub"])
    is_admin = user.get("role") == "admin"
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT user_id FROM doc_comments WHERE id=$1", _uuid_mod.UUID(comment_id))
        if not row:
            raise HTTPException(status_code=404, detail="评论不存在")
        if not is_admin and str(row["user_id"]) != str(user_uuid):
            raise HTTPException(status_code=403, detail="无权删除此评论")
        await conn.execute("DELETE FROM doc_comments WHERE id=$1", _uuid_mod.UUID(comment_id))


@app.get("/auth/me")
async def me(user=Depends(current_user)):
    pool = await get_pool()
    import uuid
    if user.get("role") == "admin":
        return {"role": "admin"}
    async with pool.acquire() as conn:
        vendor = await conn.fetchrow(VENDOR_SELECT, uuid.UUID(user["sub"]))
    if not vendor:
        raise HTTPException(status_code=404, detail="供应商信息不存在")
    return {"role": "vendor", "vendor": row_to_dict(vendor)}
